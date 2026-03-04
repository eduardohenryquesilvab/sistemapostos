from __future__ import annotations

import os
import io
import sqlite3
from datetime import datetime
from functools import wraps

from flask import Flask, render_template, request, redirect, session, url_for, abort, send_file
from werkzeug.security import generate_password_hash, check_password_hash

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, 'posto.db')

app = Flask(__name__, template_folder=os.path.join(BASE_DIR, 'templates'))
app.secret_key = os.environ.get('POSTO_SECRET_KEY', 'posto_fiel_2026')

# Roles: owner (dono) | manager (gerente)
DEFAULT_OWNER_USER = 'admin'
DEFAULT_OWNER_PASS = 'admin123'


DATABASE_URL = os.getenv("DATABASE_URL", "").strip()
DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1) if DATABASE_URL.startswith("postgres://") else DATABASE_URL
USE_POSTGRES = DATABASE_URL.lower().startswith("postgres")

class DBConn:
    """Wrapper para padronizar .execute() entre SQLite e Postgres."""
    def __init__(self, conn, is_pg: bool):
        self._conn = conn
        self.is_pg = is_pg

    def execute(self, sql: str, params=()):
        if self.is_pg:
            sql_pg = sql.replace("?", "%s")
            cur = self._conn.cursor()
            cur.execute(sql_pg, params)
            return cur
        return self._conn.execute(sql, params)

    def commit(self):
        return self._conn.commit()

    def close(self):
        return self._conn.close()

    def cursor(self):
        return self._conn.cursor()

def get_db_connection() -> DBConn:
    """Abre conexão no banco (Postgres no Railway via DATABASE_URL; SQLite local)."""
    if USE_POSTGRES:
        import psycopg2
        from psycopg2.extras import RealDictCursor
        conn = psycopg2.connect(DATABASE_URL, cursor_factory=RealDictCursor)
        return DBConn(conn, True)

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute('PRAGMA foreign_keys = ON')
    return DBConn(conn, False)
def init_db() -> None:
    """Inicializa (ou migra) o SQLite mantendo o padrão empresarial multi-posto."""
    conn = get_db_connection()
    cur = conn.cursor()
    # Se estiver em Postgres (Railway), cria/garante as tabelas no formato Postgres.
    if getattr(conn, "is_pg", False):
        # Postos
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS postos (
                id SERIAL PRIMARY KEY,
                nome_posto TEXT NOT NULL UNIQUE,
                cidade TEXT
            )
            """
        )
        # Usuários
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                role TEXT NOT NULL,
                posto_id INTEGER NULL REFERENCES postos(id),
                active INTEGER NOT NULL DEFAULT 1,
                created_at TEXT NOT NULL
            )
            """
        )
        # Colaboradores
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS colaboradores (
                id SERIAL PRIMARY KEY,
                posto_id INTEGER NOT NULL REFERENCES postos(id),
                nome TEXT NOT NULL,
                cargo TEXT,
                active INTEGER NOT NULL DEFAULT 1
            )
            """
        )
        # Estoque combustíveis
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS estoque (
                id SERIAL PRIMARY KEY,
                posto_id INTEGER NOT NULL REFERENCES postos(id),
                combustivel TEXT NOT NULL,
                litros_atuais DOUBLE PRECISION NOT NULL DEFAULT 0,
                capacidade_max DOUBLE PRECISION NOT NULL DEFAULT 0,
                UNIQUE(posto_id, combustivel)
            )
            """
        )
        # Vendas (turno)
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS vendas (
                id SERIAL PRIMARY KEY,
                posto_id INTEGER NOT NULL REFERENCES postos(id),
                data TEXT NOT NULL,
                turno TEXT,
                colaborador_id INTEGER NULL REFERENCES colaboradores(id),
                dinheiro DOUBLE PRECISION NOT NULL DEFAULT 0,
                pix DOUBLE PRECISION NOT NULL DEFAULT 0,
                debito DOUBLE PRECISION NOT NULL DEFAULT 0,
                credito DOUBLE PRECISION NOT NULL DEFAULT 0,
                litros_gasolina DOUBLE PRECISION NOT NULL DEFAULT 0,
                litros_etanol DOUBLE PRECISION NOT NULL DEFAULT 0,
                litros_diesel_s500 DOUBLE PRECISION NOT NULL DEFAULT 0,
                litros_diesel_s10 DOUBLE PRECISION NOT NULL DEFAULT 0,
                preco_gasolina DOUBLE PRECISION NOT NULL DEFAULT 0,
                preco_etanol DOUBLE PRECISION NOT NULL DEFAULT 0,
                preco_diesel_s500 DOUBLE PRECISION NOT NULL DEFAULT 0,
                preco_diesel_s10 DOUBLE PRECISION NOT NULL DEFAULT 0,
                qtd_gas INTEGER NOT NULL DEFAULT 0,
                qtd_agua INTEGER NOT NULL DEFAULT 0,
                valor_produtos DOUBLE PRECISION NOT NULL DEFAULT 0,
                notas TEXT,
                created_at TEXT NOT NULL
            )
            """
        )
        
        for old_c, new_c in [
            ("litros_gas", "litros_gasolina"), 
            ("litros_alcool", "litros_etanol"), 
            ("preco_gas", "preco_gasolina"), 
            ("preco_alcool", "preco_etanol")
        ]:
            try:
                cur.execute(f"SAVEPOINT sp_{old_c}")
                cur.execute(f"ALTER TABLE vendas RENAME COLUMN {old_c} TO {new_c}")
                cur.execute(f"RELEASE SAVEPOINT sp_{old_c}")
            except Exception:
                cur.execute(f"ROLLBACK TO SAVEPOINT sp_{old_c}")

        # Compras (NF combustíveis)
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS compras_estoque (
                id SERIAL PRIMARY KEY,
                data TEXT NOT NULL,
                posto_id INTEGER NOT NULL REFERENCES postos(id),
                combustivel TEXT NOT NULL,
                litros_comprados DOUBLE PRECISION NOT NULL,
                valor_total DOUBLE PRECISION NOT NULL,
                user_id INTEGER NULL REFERENCES users(id),
                created_at TEXT NOT NULL
            )
            """
        )
        # Despesas
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS despesas (
                id SERIAL PRIMARY KEY,
                data TEXT NOT NULL,
                posto_id INTEGER NOT NULL REFERENCES postos(id),
                categoria TEXT NOT NULL,
                descricao TEXT,
                valor DOUBLE PRECISION NOT NULL,
                forma_pagamento TEXT,
                user_id INTEGER NULL REFERENCES users(id),
                created_at TEXT NOT NULL
            )
            """
        )
        # Itens (estoque)
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS itens_estoque (
                id SERIAL PRIMARY KEY,
                posto_id INTEGER NOT NULL REFERENCES postos(id),
                categoria TEXT NOT NULL DEFAULT 'Outros',
                nome TEXT NOT NULL,
                unidade TEXT NOT NULL DEFAULT 'un',
                quantidade DOUBLE PRECISION NOT NULL DEFAULT 0,
                estoque_min DOUBLE PRECISION NOT NULL DEFAULT 0,
                custo_unit DOUBLE PRECISION NOT NULL DEFAULT 0,
                preco_venda DOUBLE PRECISION NOT NULL DEFAULT 0,
                active INTEGER NOT NULL DEFAULT 1,
                UNIQUE(posto_id, nome)
            )
            """
        )
        # Movimentações itens
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS itens_mov (
                id SERIAL PRIMARY KEY,
                data TEXT NOT NULL,
                posto_id INTEGER NOT NULL REFERENCES postos(id),
                item_id INTEGER NULL REFERENCES itens_estoque(id),
                tipo TEXT NOT NULL,
                quantidade DOUBLE PRECISION NOT NULL,
                custo_unit DOUBLE PRECISION NOT NULL DEFAULT 0,
                preco_venda DOUBLE PRECISION NOT NULL DEFAULT 0,
                observacao TEXT,
                user_id INTEGER NULL REFERENCES users(id),
                created_at TEXT NOT NULL
            )
            """
        )
        # Transferências
        cur.execute(
            """
            CREATE TABLE IF NOT EXISTS transferencias (
                id SERIAL PRIMARY KEY,
                data TEXT NOT NULL,
                origem_posto_id INTEGER NOT NULL REFERENCES postos(id),
                destino_posto_id INTEGER NOT NULL REFERENCES postos(id),
                tipo TEXT NOT NULL,
                produto TEXT NOT NULL,
                quantidade DOUBLE PRECISION NOT NULL,
                custo_unit DOUBLE PRECISION NOT NULL DEFAULT 0,
                observacao TEXT,
                user_id INTEGER NULL REFERENCES users(id),
                created_at TEXT NOT NULL
            )
            """
        )

        # Seed 5 postos
        cur.execute("SELECT 1 FROM postos LIMIT 1")
        if not cur.fetchone():
            for i in range(1, 6):
                cur.execute("INSERT INTO postos (nome_posto, cidade) VALUES (%s, %s)", (f"Posto 0{i}", ""))

        # Seed owner
        cur.execute("SELECT 1 FROM users WHERE role = %s LIMIT 1", ("owner",))
        if not cur.fetchone():
            cur.execute(
                "INSERT INTO users (username, password_hash, role, posto_id, created_at) VALUES (%s,%s,%s,%s,%s)",
                (DEFAULT_OWNER_USER, generate_password_hash(DEFAULT_OWNER_PASS), "owner", None, datetime.now().isoformat()),
            )

        # Seed tanques padrão por posto (se vazio)
        cur.execute("SELECT 1 FROM estoque LIMIT 1")
        if not cur.fetchone():
            cur.execute("SELECT id FROM postos ORDER BY id")
            postos = cur.fetchall()
            for p in postos:
                pid = p["id"] if isinstance(p, dict) else p[0]
                for comb, litros, cap in [
                    ("Gasolina Comum", 5000, 15000),
                    ("Álcool", 2000, 10000),
                    ("Diesel S500", 3000, 15000),
                    ("Diesel S10", 4000, 15000),
                ]:
                    cur.execute(
                        "INSERT INTO estoque (posto_id, combustivel, litros_atuais, capacidade_max) VALUES (%s,%s,%s,%s) ON CONFLICT (posto_id, combustivel) DO NOTHING",
                        (pid, comb, float(litros), float(cap)),
                    )

        conn.commit()
        conn.close()
        return

    def table_exists(name: str) -> bool:
        return bool(cur.execute("SELECT 1 FROM sqlite_master WHERE type='table' AND name=?", (name,)).fetchone())

    def colnames(table: str) -> list[str]:
        return [r['name'] for r in conn.execute(f"PRAGMA table_info({table})").fetchall()]

    # ------------------------
    # Postos
    # ------------------------
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS postos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome_posto TEXT NOT NULL UNIQUE,
            cidade TEXT
        )
        """
    )

    # Seed 5 postos (se vazio)
    if not cur.execute('SELECT 1 FROM postos LIMIT 1').fetchone():
        cur.executemany(
            'INSERT INTO postos (nome_posto, cidade) VALUES (?, ?)',
            [(f'Posto 0{i}', '') for i in range(1, 6)],
        )

    # ------------------------
    # Usuários
    # ------------------------
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role TEXT NOT NULL, -- owner | manager
            posto_id INTEGER,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            FOREIGN KEY(posto_id) REFERENCES postos(id)
        )
        """
    )

    # Seed: usuário dono (não exibir na UI, mas existe no banco)
    if not cur.execute('SELECT 1 FROM users WHERE role = "owner" LIMIT 1').fetchone():
        cur.execute(
            'INSERT INTO users (username, password_hash, role, posto_id, created_at) VALUES (?,?,?,?,?)',
            (
                DEFAULT_OWNER_USER,
                generate_password_hash(DEFAULT_OWNER_PASS),
                'owner',
                None,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            ),
        )

    # ------------------------
    # Operação: colaboradores, vendas, estoque de combustíveis
    # ------------------------
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS colaboradores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            posto_id INTEGER NOT NULL,
            nome TEXT NOT NULL,
            cargo TEXT,
            active INTEGER NOT NULL DEFAULT 1,
            FOREIGN KEY(posto_id) REFERENCES postos(id)
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS estoque (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            posto_id INTEGER NOT NULL,
            combustivel TEXT NOT NULL,
            litros_atuais REAL NOT NULL DEFAULT 0,
            capacidade_max REAL NOT NULL DEFAULT 0,
            UNIQUE(posto_id, combustivel),
            FOREIGN KEY(posto_id) REFERENCES postos(id)
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS vendas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            posto_id INTEGER NOT NULL,
            data TEXT NOT NULL,
            turno TEXT,
            colaborador_id INTEGER,

            dinheiro REAL NOT NULL DEFAULT 0,
            pix REAL NOT NULL DEFAULT 0,
            debito REAL NOT NULL DEFAULT 0,
            credito REAL NOT NULL DEFAULT 0,

            litros_gasolina REAL NOT NULL DEFAULT 0,
            litros_etanol REAL NOT NULL DEFAULT 0,
            litros_diesel_s500 REAL NOT NULL DEFAULT 0,
            litros_diesel_s10 REAL NOT NULL DEFAULT 0,

            preco_gasolina REAL NOT NULL DEFAULT 0,
            preco_etanol REAL NOT NULL DEFAULT 0,
            preco_diesel_s500 REAL NOT NULL DEFAULT 0,
            preco_diesel_s10 REAL NOT NULL DEFAULT 0,

            qtd_gas INTEGER NOT NULL DEFAULT 0,
            qtd_agua INTEGER NOT NULL DEFAULT 0,
            valor_produtos REAL NOT NULL DEFAULT 0,
            notas TEXT,

            created_at TEXT NOT NULL,

            FOREIGN KEY(posto_id) REFERENCES postos(id),
            FOREIGN KEY(colaborador_id) REFERENCES colaboradores(id)
        )
        """
    )

    # ------------------------
    # Compras de combustível (NF) e despesas
    # ------------------------
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS compras_estoque (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data TEXT NOT NULL,
            posto_id INTEGER NOT NULL,
            combustivel TEXT NOT NULL,
            litros_comprados REAL NOT NULL,
            valor_total REAL NOT NULL,
            user_id INTEGER,
            created_at TEXT NOT NULL,
            FOREIGN KEY(posto_id) REFERENCES postos(id),
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS despesas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data TEXT NOT NULL,
            posto_id INTEGER NOT NULL,
            categoria TEXT NOT NULL,
            descricao TEXT,
            valor REAL NOT NULL,
            forma_pagamento TEXT,
            user_id INTEGER,
            created_at TEXT NOT NULL,
            FOREIGN KEY(posto_id) REFERENCES postos(id),
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
        """
    )

    # ------------------------
    # Estoque de itens + movimentações + transferências
    # ------------------------
    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS itens_estoque (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            posto_id INTEGER NOT NULL,
            categoria TEXT NOT NULL DEFAULT 'Outros',
            nome TEXT NOT NULL,
            unidade TEXT NOT NULL DEFAULT 'un',
            quantidade REAL NOT NULL DEFAULT 0,
            estoque_min REAL NOT NULL DEFAULT 0,
            custo_unit REAL NOT NULL DEFAULT 0,
            preco_venda REAL NOT NULL DEFAULT 0,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            UNIQUE(posto_id, nome),
            FOREIGN KEY(posto_id) REFERENCES postos(id)
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS itens_mov (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data TEXT NOT NULL,
            posto_id INTEGER NOT NULL,
            item_id INTEGER,
            tipo TEXT NOT NULL, -- entrada | saida | ajuste | transferencia_in | transferencia_out
            quantidade REAL NOT NULL,
            custo_unit REAL NOT NULL DEFAULT 0,
            preco_venda REAL NOT NULL DEFAULT 0,
            ref TEXT,
            user_id INTEGER,
            created_at TEXT NOT NULL,
            FOREIGN KEY(posto_id) REFERENCES postos(id),
            FOREIGN KEY(item_id) REFERENCES itens_estoque(id),
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
        """
    )

    cur.execute(
        """
        CREATE TABLE IF NOT EXISTS transferencias (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            data TEXT NOT NULL,
            origem_posto_id INTEGER NOT NULL,
            destino_posto_id INTEGER NOT NULL,
            tipo TEXT NOT NULL, -- combustivel | item
            produto TEXT NOT NULL,
            quantidade REAL NOT NULL,
            custo_unit REAL NOT NULL DEFAULT 0,
            observacao TEXT,
            user_id INTEGER,
            created_at TEXT NOT NULL,
            FOREIGN KEY(origem_posto_id) REFERENCES postos(id),
            FOREIGN KEY(destino_posto_id) REFERENCES postos(id),
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
        """
    )

    # Compatibilidade: garantir colunas em DBs antigos
    if table_exists('vendas'):
        venda_cols = set(colnames('vendas'))
        for col, ddl in [
            ('preco_gasolina', 'REAL NOT NULL DEFAULT 0'),
            ('preco_etanol', 'REAL NOT NULL DEFAULT 0'),
            ('preco_diesel_s500', 'REAL NOT NULL DEFAULT 0'),
            ('preco_diesel_s10', 'REAL NOT NULL DEFAULT 0'),
        ]:
            if col not in venda_cols:
                cur.execute(f'ALTER TABLE vendas ADD COLUMN {col} {ddl}')

    # Seeds: combustíveis por posto
    combustiveis_padrao = [
        ('Gasolina Comum', 0, 15000),
        ('Álcool', 0, 10000),
        ('Diesel S500', 0, 15000),
        ('Diesel S10', 0, 15000),
    ]
    postos = cur.execute('SELECT id FROM postos').fetchall()
    for p in postos:
        for comb, litros, cap in combustiveis_padrao:
            cur.execute(
                'INSERT OR IGNORE INTO estoque (posto_id, combustivel, litros_atuais, capacidade_max) VALUES (?,?,?,?)',
                (p['id'], comb, litros, cap),
            )

    # Seeds: itens por posto
    itens_padrao = [
        ('Outros', 'Água', 'un', 0, 10, 0, 0),
        ('Outros', 'Gás', 'un', 0, 10, 0, 0),
    ]
    for p in postos:
        for cat, nome, un, qtd, minimo, custo, preco in itens_padrao:
            cur.execute(
                'INSERT OR IGNORE INTO itens_estoque (posto_id, categoria, nome, unidade, quantidade, estoque_min, custo_unit, preco_venda, created_at) VALUES (?,?,?,?,?,?,?,?,?)',
                (p['id'], cat, nome, un, qtd, minimo, custo, preco, datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
            )

    conn.commit()
    conn.close()

try:
    init_db()
except Exception as e:
    print(f"Erro ao inicializar db: {e}")

@app.template_filter('format_data')
def format_data(value):
    try:
        return datetime.strptime(value, '%Y-%m-%d').strftime('%d/%m/%Y')
    except Exception:
        return value


# ------------------------
# Auth helpers
# ------------------------

def current_user():
    uid = session.get('user_id')
    if not uid:
        return None
    conn = get_db_connection()
    u = conn.execute('SELECT * FROM users WHERE id = ? AND active = 1', (uid,)).fetchone()
    conn.close()
    return u


def login_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        if not session.get('user_id'):
            return redirect(url_for('login'))
        return fn(*args, **kwargs)
    return wrapper


def owner_required(fn):
    @wraps(fn)
    def wrapper(*args, **kwargs):
        u = current_user()
        if not u or u['role'] != 'owner':
            abort(403)
        return fn(*args, **kwargs)
    return wrapper


def resolve_selected_posto(u_row) -> int:
    """Owner pode selecionar via ?posto_id=. Manager é fixo no posto dele."""
    if u_row['role'] != 'owner':
        if not u_row['posto_id']:
            abort(403)
        return int(u_row['posto_id'])

    posto_qs = request.args.get('posto_id')
    conn = get_db_connection()
    if posto_qs:
        p = conn.execute('SELECT id FROM postos WHERE id = ?', (posto_qs,)).fetchone()
        if p:
            conn.close()
            return int(p['id'])
    p = conn.execute('SELECT id FROM postos ORDER BY nome_posto LIMIT 1').fetchone()
    conn.close()
    return int(p['id']) if p else 0




# ------------------------
# Estoque: alertas e estimativas
# ------------------------

FUEL_TO_VENDA_COL = {
    'Gasolina Comum': 'litros_gasolina',
    'Álcool': 'litros_etanol',
    'Diesel S500': 'litros_diesel_s500',
    'Diesel S10': 'litros_diesel_s10',
}


def calc_tank_days_left(conn, posto_id: int, combustivel: str, litros_atuais: float, lookback_days: int = 30):
    """Estima dias restantes com base no consumo médio diário (últimos N dias)."""
    col = FUEL_TO_VENDA_COL.get(combustivel)
    if not col:
        return None

    # Consumo médio diário: soma dos litros / número de dias com dados (ou lookback)
    from datetime import datetime, timedelta
    cutoff = (datetime.now() - timedelta(days=lookback_days)).strftime('%Y-%m-%d')
    
    rows = conn.execute(
        f"""
        SELECT data, SUM({col}) as litros
        FROM vendas
        WHERE posto_id = ?
          AND data >= ?
        GROUP BY data
        """,
        (posto_id, cutoff),
    ).fetchall()

    if not rows:
        return None

    total = sum(float(r['litros'] or 0) for r in rows)
    dias = len(rows)
    if dias <= 0 or total <= 0:
        return None

    media_dia = total / dias
    if media_dia <= 0:
        return None

    return round(litros_atuais / media_dia, 1)


def get_tanques_status(conn, posto_id: int, alerta_pct: float = 0.20):
    tanques = conn.execute(
        'SELECT * FROM estoque WHERE posto_id = ? ORDER BY combustivel',
        (posto_id,),
    ).fetchall()

    out = []
    for t in tanques:
        litros = float(t['litros_atuais'] or 0)
        cap = float(t['capacidade_max'] or 0)
        pct = (litros / cap) if cap > 0 else 0
        alerta = pct <= alerta_pct
        dias_left = calc_tank_days_left(conn, posto_id, t['combustivel'], litros)
        out.append({
            'id': int(t['id']),
            'combustivel': t['combustivel'],
            'litros_atuais': litros,
            'capacidade_max': cap,
            'pct': round(pct * 100, 1),
            'alerta': alerta,
            'dias_left': dias_left,
        })
    return out

# ------------------------
# AUTH (Gerencial)
# ------------------------


@app.route('/login', methods=['GET', 'POST'])
def login():
    """Login da área gerencial.

    - Não exibe credenciais padrão na UI.
    - Se o usuário já estiver logado, redireciona direto.
    """
    if session.get('user_id'):
        return redirect(url_for('gerencial'))

    erro = None
    if request.method == 'POST':
        username = (request.form.get('username') or '').strip()
        senha = request.form.get('senha') or ''

        conn = get_db_connection()
        u = conn.execute('SELECT * FROM users WHERE username = ? AND active = 1', (username,)).fetchone()
        conn.close()

        if u and check_password_hash(u['password_hash'], senha):
            session['user_id'] = int(u['id'])
            session['role'] = u['role']
            session['user_posto_id'] = int(u['posto_id']) if u['posto_id'] else None
            return redirect(url_for('gerencial'))

        erro = 'Usuário ou senha inválidos.'

    return render_template('login.html', erro=erro, app_name='EC Manager V1.0')



# ------------------------
# Fluxo FRENTISTA (Seleção do Posto -> Lançamento)
# ------------------------

@app.route('/', methods=['GET', 'POST'])
def selecionar_posto():
    """Tela simples antes do formulário para definir qual unidade está sendo lançada."""
    conn = get_db_connection()
    postos = conn.execute('SELECT * FROM postos ORDER BY nome_posto').fetchall()
    conn.close()

    if request.method == 'POST':
        posto_id = int(request.form.get('posto_id') or 0)
        if posto_id:
            session['posto_selecionado_id'] = posto_id
            return redirect(url_for('lancamento'))
    return render_template('selecionar_posto.html', postos=postos)


@app.route('/trocar-posto')
def trocar_posto():
    session.pop('posto_selecionado_id', None)
    return redirect(url_for('selecionar_posto'))


@app.route('/lancamento')
def lancamento():
    posto_id = int(session.get('posto_selecionado_id') or 0)
    if not posto_id:
        return redirect(url_for('selecionar_posto'))

    conn = get_db_connection()
    posto = conn.execute('SELECT * FROM postos WHERE id = ?', (posto_id,)).fetchone()
    colabs = conn.execute(
        'SELECT * FROM colaboradores WHERE active = 1 AND posto_id = ? ORDER BY nome',
        (posto_id,),
    ).fetchall()
    conn.close()

    return render_template(
        'index.html',
        colaboradores=colabs,
        posto=posto,
        posto_id=posto_id,
        data_hoje=datetime.now().strftime('%Y-%m-%d'),
    )


@app.route('/salvar', methods=['POST'])
def salvar():
    d = request.form
    # posto_id vem do formulário (hidden) ou da sessão
    posto_id = int(d.get('posto_id') or session.get('posto_selecionado_id') or 0)
    if not posto_id:
        return "<script>alert('Selecione a unidade (posto).'); window.location.href='/';</script>"

    def fnum(x):
        try:
            return float(str(x).replace(',', '.'))
        except Exception:
            return 0.0

    def inum(x):
        try:
            return int(float(str(x).replace(',', '.')))
        except Exception:
            return 0

    dinheiro = fnum(d.get('valor_dinheiro'))
    pix = fnum(d.get('valor_pix'))
    debito = fnum(d.get('valor_debito'))
    credito = fnum(d.get('valor_credito'))
    valor_produtos = fnum(d.get('valor_produtos'))

    # Litros vendidos
    litros_gas = fnum(d.get('litros_gas'))
    litros_alcool = fnum(d.get('litros_alcool'))
    litros_diesel_s500 = fnum(d.get('litros_diesel_s500'))
    litros_diesel_s10 = fnum(d.get('litros_diesel_s10'))

    # Itens vendidos
    qtd_gas = inum(d.get('qtd_gas'))
    qtd_agua = inum(d.get('qtd_agua'))

    # Preço/L (opcional)
    preco_gas = fnum(d.get('preco_gas'))
    preco_alcool = fnum(d.get('preco_alcool'))
    preco_diesel_s500 = fnum(d.get('preco_diesel_s500'))
    preco_diesel_s10 = fnum(d.get('preco_diesel_s10'))

    conn = get_db_connection()

    # Resolver colaborador_id pelo nome
    colaborador_nome = (d.get('colaborador') or '').strip()
    colab_row = None
    if colaborador_nome:
        colab_row = conn.execute(
            'SELECT id FROM colaboradores WHERE posto_id = ? AND nome = ? AND active = 1',
            (posto_id, colaborador_nome),
        ).fetchone()
    colaborador_id = int(colab_row['id']) if colab_row else None

    conn.execute(
        '''
        INSERT INTO vendas (
            posto_id, data, turno, colaborador_id,
            dinheiro, pix, debito, credito,
            litros_gasolina, litros_etanol, litros_diesel_s500, litros_diesel_s10,
            preco_gasolina, preco_etanol, preco_diesel_s500, preco_diesel_s10,
            qtd_gas, qtd_agua, valor_produtos, notas,
            created_at
        )
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)
        ''',
        (
            posto_id,
            d.get('data_dia'),
            d.get('turno'),
            colaborador_id,
            dinheiro, pix, debito, credito,
            litros_gas, litros_alcool, litros_diesel_s500, litros_diesel_s10,
            preco_gas, preco_alcool, preco_diesel_s500, preco_diesel_s10,
            qtd_gas,
            qtd_agua,
            valor_produtos,
            d.get('notas'),
            datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
        ),
    )

    # Baixa automática do estoque por litros
    def baixa(comb: str, litros: float):
        if litros and litros > 0:
            conn.execute(
                'UPDATE estoque SET litros_atuais = MAX(litros_atuais - ?, 0) WHERE posto_id = ? AND combustivel = ?',
                (litros, posto_id, comb),
            )

    baixa('Gasolina Comum', litros_gas)
    baixa('Álcool', litros_alcool)
    baixa('Diesel S500', litros_diesel_s500)
    baixa('Diesel S10', litros_diesel_s10)

    # Baixa automática de itens (Gás/Água) se existirem no estoque de itens
    def baixa_item(nome_item: str, qtd: int):
        if qtd and qtd > 0:
            row = conn.execute(
                'SELECT id, custo_unit, preco_venda FROM itens_estoque WHERE posto_id = ? AND nome = ? AND active = 1',
                (posto_id, nome_item),
            ).fetchone()
            if row:
                conn.execute(
                    'UPDATE itens_estoque SET quantidade = MAX(quantidade - ?, 0) WHERE id = ?',
                    (qtd, row['id']),
                )
                conn.execute(
                    'INSERT INTO itens_mov (data, posto_id, item_id, tipo, quantidade, custo_unit, preco_venda, ref, user_id, created_at) VALUES (?,?,?,?,?,?,?,?,?,?)',
                    (
                        d.get('data_dia') or datetime.now().strftime('%Y-%m-%d'),
                        posto_id,
                        row['id'],
                        'saida',
                        float(qtd),
                        float(row['custo_unit'] or 0),
                        float(row['preco_venda'] or 0),
                        'venda_turno',
                        None,
                        datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    ),
                )

    baixa_item('Água', qtd_agua)
    baixa_item('Gás', qtd_gas)

    conn.commit()
    conn.close()
    return "<script>alert('Sucesso!'); window.location.href='/lancamento';</script>"


@app.route('/gerencial')
@login_required
def gerencial():
    try:
        return _gerencial()
    except Exception as e:
        import traceback
        return f"<pre>{traceback.format_exc()}</pre>"

def _gerencial():
    u = current_user()
    conn = get_db_connection()

    postos_visiveis = conn.execute('SELECT * FROM postos ORDER BY nome_posto').fetchall()
    selected_posto_id = resolve_selected_posto(u)

    hoje = datetime.now().strftime('%Y-%m-%d')
    mes = (request.args.get('mes') or datetime.now().strftime('%Y-%m')).strip()

    vendas_hoje = conn.execute(
        """
        SELECT v.*, (v.dinheiro+v.pix+v.debito+v.credito+v.valor_produtos) as total,
               c.nome as colaborador_nome
        FROM vendas v
        LEFT JOIN colaboradores c ON c.id = v.colaborador_id
        WHERE v.data = ? AND v.posto_id = ?
        ORDER BY v.id DESC
        """,
        (hoje, selected_posto_id),
    ).fetchall()

    total_mes = conn.execute(
        """
        SELECT
            SUM(litros_gasolina) as litros_gasolina,
            SUM(litros_etanol) as litros_etanol,
            SUM(litros_diesel_s500) as litros_diesel_s500,
            SUM(litros_diesel_s10) as litros_diesel_s10,
            SUM(qtd_gas) as gas,
            SUM(qtd_agua) as agua,
            SUM(valor_produtos) as lubs,
            SUM(dinheiro+pix+debito+credito+valor_produtos) as financeiro,
            SUM(dinheiro) as dinheiro,
            SUM(pix) as pix,
            SUM(debito) as debito,
            SUM(credito) as credito
        FROM vendas
        WHERE data LIKE ? AND posto_id = ?
        """,
        (f'{mes}%', selected_posto_id),
    ).fetchone()

    produtividade = conn.execute(
        """
        SELECT COALESCE(c.nome, 'Sem colaborador') as colaborador,
               SUM(v.dinheiro+v.pix+v.debito+v.credito+v.valor_produtos) as total
        FROM vendas v
        LEFT JOIN colaboradores c ON c.id = v.colaborador_id
        WHERE v.data LIKE ? AND v.posto_id = ?
        GROUP BY COALESCE(c.nome, 'Sem colaborador')
        ORDER BY total DESC
        """,
        (f'{mes}%', selected_posto_id),
    ).fetchall()

    tanques = conn.execute(
        'SELECT * FROM estoque WHERE posto_id = ? ORDER BY combustivel',
        (selected_posto_id,),
    ).fetchall()



    # Alertas: tanques abaixo de 20% e estimativa de dias restantes
    tanques_status = get_tanques_status(conn, selected_posto_id, alerta_pct=0.20)

    # Alertas: itens abaixo do mínimo
    try:
        itens_alerta = conn.execute(
            "SELECT * FROM itens_estoque WHERE posto_id = ? AND active = 1 AND quantidade <= estoque_min ORDER BY (estoque_min - quantidade) DESC",
            (selected_posto_id,),
        ).fetchall()
    except Exception:
        itens_alerta = []

    despesas_mes_val = conn.execute(
        'SELECT COALESCE(SUM(valor),0) as total FROM despesas WHERE data LIKE ? AND posto_id = ?',
        (f'{mes}%', selected_posto_id),
    ).fetchone()
    despesas_mes = float(despesas_mes_val['total'] or 0) if despesas_mes_val else 0.0

    # Receita estimada de combustíveis (se informado preço por litro no lançamento)
    receita_comb_val = conn.execute(
        '''
        SELECT COALESCE(SUM(
            litros_gasolina*preco_gasolina +
            litros_etanol*preco_etanol +
            litros_diesel_s500*preco_diesel_s500 +
            litros_diesel_s10*preco_diesel_s10
        ),0) as total
        FROM vendas
        WHERE data LIKE ? AND posto_id = ?
        ''',
        (f'{mes}%', selected_posto_id),
    ).fetchone()
    receita_comb_mes = float(receita_comb_val['total'] or 0) if receita_comb_val else 0.0

    # Custo médio por litro (ponderado) com base nas NFs lançadas
    custos = conn.execute(
        '''
        SELECT combustivel,
               COALESCE(SUM(valor_total) / NULLIF(SUM(litros_comprados),0), 0) as custo_medio
        FROM compras_estoque
        WHERE posto_id = ?
        GROUP BY combustivel
        ''',
        (selected_posto_id,),
    ).fetchall()
    custo_medio = {r['combustivel']: float(r['custo_medio'] or 0) for r in custos}

    litros_g = float(total_mes['litros_gasolina'] or 0) if total_mes else 0.0
    litros_e = float(total_mes['litros_etanol'] or 0) if total_mes else 0.0
    litros_500 = float(total_mes['litros_diesel_s500'] or 0) if total_mes else 0.0
    litros_10 = float(total_mes['litros_diesel_s10'] or 0) if total_mes else 0.0

    custo_comb_mes = (
        litros_g * custo_medio.get('Gasolina Comum', 0) +
        litros_e * custo_medio.get('Álcool', 0) +
        litros_500 * custo_medio.get('Diesel S500', 0) +
        litros_10 * custo_medio.get('Diesel S10', 0)
    )

    lucro_bruto_comb_mes = receita_comb_mes - custo_comb_mes
    financeiro_mes = float(total_mes['financeiro'] or 0) if total_mes else 0.0
    lucro_liquido_estimado = financeiro_mes - despesas_mes - custo_comb_mes
    conn.close()

    return render_template(
        'gerencial.html',
        user=u,
        postos=postos_visiveis,
        posto_id=selected_posto_id,
        hoje=hoje,
        mes=mes,
        vendas=vendas_hoje,
        total_mes=total_mes,
        produtividade=produtividade,
        tanques=tanques,
        tanques_status=tanques_status,
        itens_alerta=itens_alerta,
        despesas_mes=despesas_mes,
        receita_comb_mes=receita_comb_mes,
        custo_comb_mes=custo_comb_mes,
        lucro_bruto_comb_mes=lucro_bruto_comb_mes,
        lucro_liquido_estimado=lucro_liquido_estimado,
    )




# ------------------------
# Export (Excel/PDF)
# ------------------------

def _fetch_report_data(conn: sqlite3.Connection, posto_id: int, mes: str) -> dict:
    vendas = conn.execute(
        '''
        SELECT v.*, p.nome_posto,
               COALESCE(c.nome, '') as colaborador_nome,
               (v.dinheiro+v.pix+v.debito+v.credito+v.valor_produtos) as total
        FROM vendas v
        JOIN postos p ON p.id = v.posto_id
        LEFT JOIN colaboradores c ON c.id = v.colaborador_id
        WHERE v.data LIKE ? AND v.posto_id = ?
        ORDER BY v.data ASC, v.id ASC
        ''',
        (f'{mes}%', posto_id),
    ).fetchall()

    despesas = conn.execute(
        '''
        SELECT d.*, p.nome_posto
        FROM despesas d
        JOIN postos p ON p.id = d.posto_id
        WHERE d.data LIKE ? AND d.posto_id = ?
        ORDER BY d.data ASC, d.id ASC
        ''',
        (f'{mes}%', posto_id),
    ).fetchall()

    compras = conn.execute(
        '''
        SELECT c.*, p.nome_posto
        FROM compras_estoque c
        JOIN postos p ON p.id = c.posto_id
        WHERE c.data LIKE ? AND c.posto_id = ?
        ORDER BY c.data ASC, c.id ASC
        ''',
        (f'{mes}%', posto_id),
    ).fetchall()

    resumo = conn.execute(
        '''
        SELECT
            COALESCE(SUM(v.dinheiro+v.pix+v.debito+v.credito+v.valor_produtos),0) as financeiro,
            COALESCE(SUM(v.dinheiro),0) as dinheiro,
            COALESCE(SUM(v.pix),0) as pix,
            COALESCE(SUM(v.debito),0) as debito,
            COALESCE(SUM(v.credito),0) as credito,

            COALESCE(SUM(v.litros_gasolina),0) as litros_gasolina,
            COALESCE(SUM(v.litros_etanol),0) as litros_etanol,
            COALESCE(SUM(v.litros_diesel_s500),0) as litros_diesel_s500,
            COALESCE(SUM(v.litros_diesel_s10),0) as litros_diesel_s10,

            COALESCE(SUM(v.litros_gasolina*v.preco_gasolina + v.litros_etanol*v.preco_etanol +
                         v.litros_diesel_s500*v.preco_diesel_s500 + v.litros_diesel_s10*v.preco_diesel_s10),0) as receita_comb

        FROM vendas v
        WHERE v.data LIKE ? AND v.posto_id = ?
        ''',
        (f'{mes}%', posto_id),
    ).fetchone()

    total_despesas = conn.execute(
        'SELECT COALESCE(SUM(valor),0) as total FROM despesas WHERE data LIKE ? AND posto_id = ?',
        (f'{mes}%', posto_id),
    ).fetchone()['total']

    custos = conn.execute(
        '''
        SELECT combustivel,
               COALESCE(SUM(valor_total) / NULLIF(SUM(litros_comprados),0), 0) as custo_medio
        FROM compras_estoque
        WHERE posto_id = ?
        GROUP BY combustivel
        ''',
        (posto_id,),
    ).fetchall()
    custo_medio = {r['combustivel']: float(r['custo_medio'] or 0) for r in custos}

    custo_comb = (
        float(resumo['litros_gasolina'] or 0) * custo_medio.get('Gasolina Comum', 0) +
        float(resumo['litros_etanol'] or 0) * custo_medio.get('Álcool', 0) +
        float(resumo['litros_diesel_s500'] or 0) * custo_medio.get('Diesel S500', 0) +
        float(resumo['litros_diesel_s10'] or 0) * custo_medio.get('Diesel S10', 0)
    )

    return {
        'vendas': vendas,
        'despesas': despesas,
        'compras': compras,
        'resumo': resumo,
        'total_despesas': float(total_despesas or 0),
        'custo_medio': custo_medio,
        'custo_comb': float(custo_comb or 0),
    }


@app.route('/gerencial/export/excel')
@login_required
def export_excel():
    u = current_user()
    posto_id = resolve_selected_posto(u)
    mes = (request.args.get('mes') or datetime.now().strftime('%Y-%m')).strip()

    conn = get_db_connection()
    posto = conn.execute('SELECT nome_posto FROM postos WHERE id = ?', (posto_id,)).fetchone()
    posto_nome = (posto['nome_posto'] if posto else f'Posto_{posto_id}').replace(' ', '_')

    data = _fetch_report_data(conn, posto_id, mes)
    conn.close()

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from openpyxl.utils import get_column_letter
    except Exception as e:
        return "<script>alert('Exportação Excel indisponível: instale openpyxl (pip install openpyxl).'); window.history.back();</script>"


    wb = Workbook()
    ws = wb.active
    ws.title = 'Resumo'

    header_fill = PatternFill('solid', fgColor='0B4DB3')
    header_font = Font(color='FFFFFF', bold=True)

    def write_kv(row, key, val):
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=val)
        ws.cell(row=row, column=1).font = Font(bold=True)

    ws['A1'] = 'EC Manager V1.0 — Relatório Mensal'
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:D1')

    write_kv(3, 'Mês', mes)
    write_kv(4, 'Unidade', posto_nome.replace('_', ' '))

    r = data['resumo']
    write_kv(6, 'Financeiro (R$)', float(r['financeiro'] or 0))
    write_kv(7, 'Dinheiro (R$)', float(r['dinheiro'] or 0))
    write_kv(8, 'PIX (R$)', float(r['pix'] or 0))
    write_kv(9, 'Débito (R$)', float(r['debito'] or 0))
    write_kv(10, 'Crédito (R$)', float(r['credito'] or 0))

    write_kv(12, 'Litros Gasolina', float(r['litros_gasolina'] or 0))
    write_kv(13, 'Litros Álcool', float(r['litros_etanol'] or 0))
    write_kv(14, 'Litros Diesel S500', float(r['litros_diesel_s500'] or 0))
    write_kv(15, 'Litros Diesel S10', float(r['litros_diesel_s10'] or 0))

    write_kv(17, 'Receita Combustível (estimada)', float(r['receita_comb'] or 0))
    write_kv(18, 'Custo Combustível (estimado)', float(data['custo_comb'] or 0))
    write_kv(19, 'Despesas (R$)', float(data['total_despesas'] or 0))
    write_kv(20, 'Lucro Líquido (estimado)', float(r['financeiro'] or 0) - float(data['total_despesas'] or 0) - float(data['custo_comb'] or 0))

    ws.column_dimensions['A'].width = 34
    ws.column_dimensions['B'].width = 22

    # Sheet Vendas
    ws_v = wb.create_sheet('Vendas')
    headers = [
        'Data', 'Turno', 'Colaborador', 'Dinheiro', 'PIX', 'Débito', 'Crédito', 'Outros Produtos',
        'L Gasolina', 'Preço Gas', 'L Álcool', 'Preço Álc', 'L Diesel S500', 'Preço S500', 'L Diesel S10', 'Preço S10',
        'Total', 'Notas'
    ]
    ws_v.append(headers)
    for col in range(1, len(headers)+1):
        c = ws_v.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')

    for v in data['vendas']:
        ws_v.append([
            v['data'], v['turno'], v['colaborador_nome'],
            float(v['dinheiro'] or 0), float(v['pix'] or 0), float(v['debito'] or 0), float(v['credito'] or 0), float(v['valor_produtos'] or 0),
            float(v['litros_gasolina'] or 0), float(v['preco_gasolina'] or 0),
            float(v['litros_etanol'] or 0), float(v['preco_etanol'] or 0),
            float(v['litros_diesel_s500'] or 0), float(v['preco_diesel_s500'] or 0),
            float(v['litros_diesel_s10'] or 0), float(v['preco_diesel_s10'] or 0),
            float(v['total'] or 0),
            v['notas'] or ''
        ])

    # Sheet Despesas
    ws_d = wb.create_sheet('Despesas')
    d_headers = ['Data', 'Categoria', 'Descrição', 'Valor', 'Forma Pagamento']
    ws_d.append(d_headers)
    for col in range(1, len(d_headers)+1):
        c = ws_d.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')
    for d in data['despesas']:
        ws_d.append([d['data'], d['categoria'], d['descricao'] or '', float(d['valor'] or 0), d['forma_pagamento'] or ''])

    # Sheet NFs
    ws_c = wb.create_sheet('NFs (Combustível)')
    c_headers = ['Data', 'Combustível', 'Litros', 'Valor Total', 'Custo por Litro']
    ws_c.append(c_headers)
    for col in range(1, len(c_headers)+1):
        c = ws_c.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal='center')
    for c in data['compras']:
        litros = float(c['litros_comprados'] or 0)
        valor = float(c['valor_total'] or 0)
        custo_l = (valor / litros) if litros else 0
        ws_c.append([c['data'], c['combustivel'], litros, valor, custo_l])

    # Auto width
    for wsx in [ws_v, ws_d, ws_c]:
        for col in range(1, wsx.max_column + 1):
            max_len = 0
            col_letter = get_column_letter(col)
            for row in range(1, min(wsx.max_row, 400) + 1):
                val = wsx.cell(row=row, column=col).value
                if val is None:
                    continue
                max_len = max(max_len, len(str(val)))
            wsx.column_dimensions[col_letter].width = min(max(12, max_len + 2), 44)

    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)

    filename = f'EC_Manager_{posto_nome}_{mes}.xlsx'
    return send_file(
        bio,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@app.route('/gerencial/export/pdf')
@login_required
def export_pdf():
    u = current_user()
    posto_id = resolve_selected_posto(u)
    mes = (request.args.get('mes') or datetime.now().strftime('%Y-%m')).strip()

    conn = get_db_connection()
    posto = conn.execute('SELECT nome_posto FROM postos WHERE id = ?', (posto_id,)).fetchone()
    posto_nome = (posto['nome_posto'] if posto else f'Posto {posto_id}')
    data = _fetch_report_data(conn, posto_id, mes)
    conn.close()

    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet
    except Exception:
        return "<script>alert('Exportação PDF indisponível: instale reportlab (pip install reportlab).'); window.history.back();</script>"


    buff = io.BytesIO()
    doc = SimpleDocTemplate(buff, pagesize=A4, topMargin=36, bottomMargin=36, leftMargin=28, rightMargin=28)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(f"<b>EC Manager V1.0</b> — Relatório Mensal ({mes})", styles['Title']))
    story.append(Paragraph(f"Unidade: <b>{posto_nome}</b>", styles['Normal']))
    story.append(Spacer(1, 12))

    r = data['resumo']
    resumo_tbl = [
        ['Financeiro (R$)', f"{float(r['financeiro'] or 0):.2f}"],
        ['Despesas (R$)', f"{float(data['total_despesas'] or 0):.2f}"],
        ['Custo Combustível (R$)', f"{float(data['custo_comb'] or 0):.2f}"],
        ['Receita Combustível (R$)', f"{float(r['receita_comb'] or 0):.2f}"],
        ['Lucro Líquido (estim.) (R$)', f"{(float(r['financeiro'] or 0) - float(data['total_despesas'] or 0) - float(data['custo_comb'] or 0)):.2f}"],
    ]
    t = Table(resumo_tbl, colWidths=[220, 120])
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#0B4DB3')),
        ('TEXTCOLOR',(0,0),(-1,0),colors.white),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('GRID',(0,0),(-1,-1),0.25,colors.grey),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.whitesmoke, colors.lightgrey]),
        ('PADDING',(0,0),(-1,-1),6),
    ]))
    story.append(Paragraph("<b>Resumo</b>", styles['Heading2']))
    story.append(t)
    story.append(Spacer(1, 14))

    # Últimas vendas (até 25 linhas)
    story.append(Paragraph("<b>Vendas (amostra)</b>", styles['Heading2']))
    v_rows = [['Data','Turno','Colaborador','Total (R$)']]
    for v in data['vendas'][-25:]:
        v_rows.append([v['data'], v['turno'] or '', v['colaborador_nome'] or '', f"{float(v['total'] or 0):.2f}"])
    tv = Table(v_rows, colWidths=[70, 60, 170, 80])
    tv.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#F7B500')),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('GRID',(0,0),(-1,-1),0.25,colors.grey),
        ('PADDING',(0,0),(-1,-1),4),
    ]))
    story.append(tv)
    story.append(Spacer(1, 14))

    # Despesas (até 25 linhas)
    story.append(Paragraph("<b>Despesas (amostra)</b>", styles['Heading2']))
    d_rows = [['Data','Categoria','Valor (R$)','Descrição']]
    for d in data['despesas'][-25:]:
        d_rows.append([d['data'], d['categoria'], f"{float(d['valor'] or 0):.2f}", (d['descricao'] or '')[:50]])
    td = Table(d_rows, colWidths=[70, 110, 80, 120])
    td.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#F7B500')),
        ('FONTNAME',(0,0),(-1,0),'Helvetica-Bold'),
        ('GRID',(0,0),(-1,-1),0.25,colors.grey),
        ('PADDING',(0,0),(-1,-1),4),
    ]))
    story.append(td)

    doc.build(story)
    buff.seek(0)

    filename = f"EC_Manager_{posto_nome.replace(' ','_')}_{mes}.pdf"
    return send_file(buff, as_attachment=True, download_name=filename, mimetype='application/pdf')



# ------------------------
# Estoque (Gerencial)
# ------------------------

@app.route('/gerencial/estoque', methods=['GET', 'POST'])
@login_required
def estoque_adm():
    u = current_user()
    selected_posto_id = resolve_selected_posto(u)

    conn = get_db_connection()
    if request.method == 'POST':
        combustivel = request.form.get('combustivel')
        tipo = request.form.get('tipo_operacao')
        valor = float(request.form.get('quantidade') or 0)
        if tipo == 'entrada':
            conn.execute(
                'UPDATE estoque SET litros_atuais = litros_atuais + ? WHERE posto_id = ? AND combustivel = ?',
                (valor, selected_posto_id, combustivel),
            )
        elif tipo == 'capacidade':
            conn.execute(
                'UPDATE estoque SET capacidade_max = ? WHERE posto_id = ? AND combustivel = ?',
                (valor, selected_posto_id, combustivel),
            )
        else:
            conn.execute(
                'UPDATE estoque SET litros_atuais = ? WHERE posto_id = ? AND combustivel = ?',
                (valor, selected_posto_id, combustivel),
            )
        conn.commit()
        conn.close()
        return redirect(url_for('estoque_adm', posto_id=selected_posto_id))

    tanques = conn.execute(
        'SELECT * FROM estoque WHERE posto_id = ? ORDER BY combustivel',
        (selected_posto_id,),
    ).fetchall()
    tanques_status = get_tanques_status(conn, selected_posto_id, alerta_pct=0.20)
    postos = conn.execute('SELECT * FROM postos ORDER BY nome_posto').fetchall()
    conn.close()

    return render_template('estoque_gerencial.html', tanques=tanques, tanques_status=tanques_status, postos=postos, posto_id=selected_posto_id, user=u)


# ------------------------
# Entradas NF (Gerencial)
# ------------------------

@app.route('/gerencial/combustiveis', methods=['GET', 'POST'])
@login_required
def combustiveis_nf():
    u = current_user()
    selected_posto_id = resolve_selected_posto(u)

    conn = get_db_connection()

    if request.method == 'POST':
        combustivel = request.form.get('combustivel')
        litros = float(request.form.get('litros') or 0)
        valor_total = float(request.form.get('valor_total') or 0)

        conn.execute(
            """
            INSERT INTO compras_estoque (data, posto_id, combustivel, litros_comprados, valor_total, user_id, created_at)
            VALUES (?,?,?,?,?,?,?)
            """,
            (
                datetime.now().strftime('%Y-%m-%d'),
                selected_posto_id,
                combustivel,
                litros,
                valor_total,
                int(u['id']) if u else None,
                datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            ),
        )

        conn.execute(
            'UPDATE estoque SET litros_atuais = litros_atuais + ? WHERE posto_id = ? AND combustivel = ?',
            (litros, selected_posto_id, combustivel),
        )
        conn.commit()
        conn.close()
        return redirect(url_for('combustiveis_nf', posto_id=selected_posto_id))

    tanques = conn.execute(
        'SELECT * FROM estoque WHERE posto_id = ? ORDER BY combustivel',
        (selected_posto_id,),
    ).fetchall()

    compras = conn.execute(
        'SELECT * FROM compras_estoque WHERE posto_id = ? ORDER BY id DESC LIMIT 200',
        (selected_posto_id,),
    ).fetchall()

    custos = conn.execute(
        '''
        SELECT combustivel,
               COALESCE(SUM(valor_total) / NULLIF(SUM(litros_comprados),0), 0) as custo_medio
        FROM compras_estoque
        WHERE posto_id = ?
        GROUP BY combustivel
        ''',
        (selected_posto_id,),
    ).fetchall()
    custo_medio = {r['combustivel']: float(r['custo_medio'] or 0) for r in custos}

    # Status dos tanques (alerta em 20%)
    tanques_status = get_tanques_status(conn, selected_posto_id, alerta_pct=0.20)

    # Alertas: itens abaixo do mínimo
    try:
        itens_alerta = conn.execute(
            "SELECT * FROM itens_estoque WHERE posto_id = ? AND active = 1 AND quantidade <= estoque_min ORDER BY (estoque_min - quantidade) DESC",
            (selected_posto_id,),
        ).fetchall()
    except Exception:
        itens_alerta = []

    postos = conn.execute('SELECT * FROM postos ORDER BY nome_posto').fetchall()
    conn.close()

    return render_template(
        'combustiveis.html',
        tanques=tanques,
        tanques_status=tanques_status,
        itens_alerta=itens_alerta,
        compras=compras,
        postos=postos,
        posto_id=selected_posto_id,
        user=u,
        custo_medio=custo_medio,
    )



# ------------------------
# Despesas (Gerencial)
# ------------------------

@app.route('/gerencial/despesas', methods=['GET', 'POST'])
@login_required
def despesas():
    u = current_user()
    selected_posto_id = resolve_selected_posto(u)
    conn = get_db_connection()

    if request.method == 'POST':
        data = request.form.get('data') or datetime.now().strftime('%Y-%m-%d')
        categoria = (request.form.get('categoria') or '').strip()
        descricao = (request.form.get('descricao') or '').strip()
        forma = (request.form.get('forma_pagamento') or '').strip()
        try:
            valor = float(request.form.get('valor') or 0)
        except Exception:
            valor = 0.0

        if categoria and valor > 0:
            conn.execute(
                '''
                INSERT INTO despesas (data, posto_id, categoria, descricao, valor, forma_pagamento, user_id, created_at)
                VALUES (?,?,?,?,?,?,?,?)
                ''',
                (
                    data,
                    selected_posto_id,
                    categoria,
                    descricao,
                    valor,
                    forma,
                    int(u['id']) if u else None,
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                ),
            )
            conn.commit()
        return redirect(url_for('despesas', posto_id=selected_posto_id))

    mes = (request.args.get('mes') or datetime.now().strftime('%Y-%m')).strip()

    despesas_mes = conn.execute(
        'SELECT COALESCE(SUM(valor),0) as total FROM despesas WHERE data LIKE ? AND posto_id = ?',
        (f'{mes}%', selected_posto_id),
    ).fetchone()['total']

    itens = conn.execute(
        'SELECT * FROM despesas WHERE data LIKE ? AND posto_id = ? ORDER BY data DESC, id DESC LIMIT 300',
        (f'{mes}%', selected_posto_id),
    ).fetchall()

    postos = conn.execute('SELECT * FROM postos ORDER BY nome_posto').fetchall()
    conn.close()

    return render_template(
        'despesas.html',
        user=u,
        postos=postos,
        posto_id=selected_posto_id,
        mes=mes,
        despesas_mes=despesas_mes,
        itens=itens,
        hoje=datetime.now().strftime('%Y-%m-%d'),
    )






# ------------------------
# Itens (Estoque de produtos) - Gerencial
# ------------------------

@app.route('/gerencial/itens', methods=['GET', 'POST'])
@login_required
def itens():
    u = current_user()
    selected_posto_id = resolve_selected_posto(u)
    conn = get_db_connection()

    if request.method == 'POST':
        action = request.form.get('action') or ''
        data_mov = request.form.get('data') or datetime.now().strftime('%Y-%m-%d')

        def fnum(x):
            try:
                return float(x)
            except Exception:
                return 0.0

        if action == 'create':
            nome = (request.form.get('nome') or '').strip()
            categoria = (request.form.get('categoria') or 'Outros').strip()
            unidade = (request.form.get('unidade') or 'un').strip()
            estoque_min = fnum(request.form.get('estoque_min'))
            custo_unit = fnum(request.form.get('custo_unit'))
            preco_venda = fnum(request.form.get('preco_venda'))
            if nome:
                conn.execute(
                    "INSERT OR IGNORE INTO itens_estoque (posto_id, categoria, nome, unidade, quantidade, estoque_min, custo_unit, preco_venda, created_at) VALUES (?,?,?,?,?,?,?,?,?)",
                    (selected_posto_id, categoria, nome, unidade, 0, estoque_min, custo_unit, preco_venda, datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                )
                conn.commit()

        elif action == 'move':
            item_id = int(request.form.get('item_id') or 0)
            tipo = request.form.get('tipo') or 'entrada'
            qtd = fnum(request.form.get('quantidade'))
            if item_id and qtd >= 0:
                row = conn.execute('SELECT * FROM itens_estoque WHERE id = ? AND posto_id = ?', (item_id, selected_posto_id)).fetchone()
                if row:
                    if tipo == 'entrada':
                        conn.execute('UPDATE itens_estoque SET quantidade = quantidade + ? WHERE id = ?', (qtd, item_id))
                        mov_tipo = 'entrada'
                    elif tipo == 'saida':
                        conn.execute('UPDATE itens_estoque SET quantidade = MAX(quantidade - ?, 0) WHERE id = ?', (qtd, item_id))
                        mov_tipo = 'saida'
                    else:
                        conn.execute('UPDATE itens_estoque SET quantidade = ? WHERE id = ?', (qtd, item_id))
                        mov_tipo = 'ajuste'

                    conn.execute(
                        'INSERT INTO itens_mov (data, posto_id, item_id, tipo, quantidade, custo_unit, preco_venda, ref, user_id, created_at) VALUES (?,?,?,?,?,?,?,?,?,?)',
                        (data_mov, selected_posto_id, item_id, mov_tipo, float(qtd), float(row['custo_unit'] or 0), float(row['preco_venda'] or 0), 'manual', int(u['id']) if u else None, datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                    )
                    conn.commit()

        elif action == 'update':
            item_id = int(request.form.get('item_id') or 0)
            if item_id:
                categoria = (request.form.get('categoria') or 'Outros').strip()
                unidade = (request.form.get('unidade') or 'un').strip()
                estoque_min = fnum(request.form.get('estoque_min'))
                custo_unit = fnum(request.form.get('custo_unit'))
                preco_venda = fnum(request.form.get('preco_venda'))
                conn.execute(
                    'UPDATE itens_estoque SET categoria=?, unidade=?, estoque_min=?, custo_unit=?, preco_venda=? WHERE id=? AND posto_id=?',
                    (categoria, unidade, estoque_min, custo_unit, preco_venda, item_id, selected_posto_id),
                )
                conn.commit()

        conn.close()
        return redirect(url_for('itens', posto_id=selected_posto_id))

    itens = conn.execute('SELECT * FROM itens_estoque WHERE posto_id = ? AND active = 1 ORDER BY categoria, nome', (selected_posto_id,)).fetchall()
    alerts = conn.execute('SELECT * FROM itens_estoque WHERE posto_id = ? AND active = 1 AND quantidade <= estoque_min ORDER BY (estoque_min - quantidade) DESC', (selected_posto_id,)).fetchall()
    postos = conn.execute('SELECT * FROM postos ORDER BY nome_posto').fetchall()
    conn.close()

    return render_template('itens.html', user=u, postos=postos, posto_id=selected_posto_id, itens=itens, alerts=alerts, hoje=datetime.now().strftime('%Y-%m-%d'))


# ------------------------
# Transferências entre postos - Gerencial
# ------------------------

@app.route('/gerencial/transferencias', methods=['GET', 'POST'])
@login_required
def transferencias():
    u = current_user()
    conn = get_db_connection()

    def fnum(x):
        try:
            return float(x)
        except Exception:
            return 0.0

    if request.method == 'POST':
        data = request.form.get('data') or datetime.now().strftime('%Y-%m-%d')
        tipo = request.form.get('tipo') or 'combustivel'
        origem = int(request.form.get('origem_posto_id') or 0)
        destino = int(request.form.get('destino_posto_id') or 0)
        produto = (request.form.get('produto') or '').strip()
        quantidade = fnum(request.form.get('quantidade'))
        obs = request.form.get('observacao')

        if u and u['role'] != 'owner':
            if not u['posto_id'] or int(u['posto_id']) != origem:
                conn.close()
                abort(403)

        if origem and destino and origem != destino and produto and quantidade > 0:
            if tipo == 'combustivel':
                conn.execute('UPDATE estoque SET litros_atuais = MAX(litros_atuais - ?, 0) WHERE posto_id = ? AND combustivel = ?', (quantidade, origem, produto))
                conn.execute('UPDATE estoque SET litros_atuais = litros_atuais + ? WHERE posto_id = ? AND combustivel = ?', (quantidade, destino, produto))
                item_origem = None
            else:
                item_origem = conn.execute('SELECT id, custo_unit, preco_venda FROM itens_estoque WHERE posto_id = ? AND nome = ? AND active = 1', (origem, produto)).fetchone()
                if item_origem:
                    conn.execute('UPDATE itens_estoque SET quantidade = MAX(quantidade - ?, 0) WHERE id = ?', (quantidade, item_origem['id']))
                    conn.execute('INSERT INTO itens_mov (data, posto_id, item_id, tipo, quantidade, custo_unit, preco_venda, ref, user_id, created_at) VALUES (?,?,?,?,?,?,?,?,?,?)',
                                (data, origem, item_origem['id'], 'transferencia_out', quantidade, float(item_origem['custo_unit'] or 0), float(item_origem['preco_venda'] or 0), f'transf:{destino}', int(u['id']) if u else None, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))

                item_dest = conn.execute('SELECT id FROM itens_estoque WHERE posto_id = ? AND nome = ? AND active = 1', (destino, produto)).fetchone()
                if not item_dest:
                    custo = float(item_origem['custo_unit'] or 0) if item_origem else 0
                    preco = float(item_origem['preco_venda'] or 0) if item_origem else 0
                    conn.execute('INSERT OR IGNORE INTO itens_estoque (posto_id, categoria, nome, unidade, quantidade, estoque_min, custo_unit, preco_venda, created_at) VALUES (?,?,?,?,?,?,?,?,?)',
                                (destino, 'Outros', produto, 'un', 0, 0, custo, preco, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
                    item_dest = conn.execute('SELECT id FROM itens_estoque WHERE posto_id = ? AND nome = ? AND active = 1', (destino, produto)).fetchone()
                if item_dest:
                    conn.execute('UPDATE itens_estoque SET quantidade = quantidade + ? WHERE id = ?', (quantidade, item_dest['id']))
                    conn.execute('INSERT INTO itens_mov (data, posto_id, item_id, tipo, quantidade, custo_unit, preco_venda, ref, user_id, created_at) VALUES (?,?,?,?,?,?,?,?,?,?)',
                                (data, destino, item_dest['id'], 'transferencia_in', quantidade, float(item_origem['custo_unit'] or 0) if item_origem else 0, float(item_origem['preco_venda'] or 0) if item_origem else 0, f'transf:{origem}', int(u['id']) if u else None, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))

            custo_unit = 0
            if tipo == 'combustivel':
                row = conn.execute('SELECT COALESCE(SUM(valor_total) / NULLIF(SUM(litros_comprados),0), 0) as cm FROM compras_estoque WHERE posto_id = ? AND combustivel = ?', (origem, produto)).fetchone()
                custo_unit = float(row['cm'] or 0) if row else 0
            else:
                if item_origem:
                    custo_unit = float(item_origem['custo_unit'] or 0)

            conn.execute('INSERT INTO transferencias (data, origem_posto_id, destino_posto_id, tipo, produto, quantidade, custo_unit, observacao, user_id, created_at) VALUES (?,?,?,?,?,?,?,?,?,?)',
                        (data, origem, destino, tipo, produto, quantidade, custo_unit, obs, int(u['id']) if u else None, datetime.now().strftime('%Y-%m-%d %H:%M:%S')))
            conn.commit()

        conn.close()
        return redirect(url_for('transferencias'))

    postos = conn.execute('SELECT * FROM postos ORDER BY nome_posto').fetchall()
    combustiveis = conn.execute('SELECT DISTINCT combustivel as nome FROM estoque ORDER BY combustivel').fetchall()
    itens_distintos = conn.execute('SELECT DISTINCT nome FROM itens_estoque WHERE active = 1 ORDER BY nome').fetchall()
    historico = conn.execute(
        "SELECT t.*, po.nome_posto as origem_nome, pd.nome_posto as destino_nome FROM transferencias t LEFT JOIN postos po ON po.id = t.origem_posto_id LEFT JOIN postos pd ON pd.id = t.destino_posto_id ORDER BY t.id DESC LIMIT 200"
    ).fetchall()
    conn.close()

    return render_template('transferencias.html', user=u, postos=postos, combustiveis=combustiveis, itens_distintos=itens_distintos, historico=historico, hoje=datetime.now().strftime('%Y-%m-%d'))

# ------------------------
# Equipe (Gerencial)
# ------------------------

@app.route('/gerencial/equipe', methods=['GET', 'POST'])
@login_required
def equipe():
    u = current_user()
    selected_posto_id = resolve_selected_posto(u)

    conn = get_db_connection()
    if request.method == 'POST':
        conn.execute(
            'INSERT INTO colaboradores (posto_id, nome, cargo) VALUES (?,?,?)',
            (selected_posto_id, request.form.get('nome'), request.form.get('cargo')),
        )
        conn.commit()

    colabs = conn.execute(
        'SELECT * FROM colaboradores WHERE posto_id = ? AND active = 1 ORDER BY nome',
        (selected_posto_id,),
    ).fetchall()
    postos = conn.execute('SELECT * FROM postos ORDER BY nome_posto').fetchall()
    conn.close()

    return render_template('equipe.html', colaboradores=colabs, postos=postos, posto_id=selected_posto_id, user=u)


# ------------------------
# Postos (Gerencial) - owner only
# ------------------------

@app.route('/gerencial/postos', methods=['GET', 'POST'])
@login_required
@owner_required
def postos():
    conn = get_db_connection()
    if request.method == 'POST':
        nome = (request.form.get('nome_posto') or '').strip()
        cidade = (request.form.get('cidade') or '').strip()
        if nome:
            conn.execute('INSERT INTO postos (nome_posto, cidade) VALUES (?, ?)', (nome, cidade))
            conn.commit()
    lista = conn.execute('SELECT * FROM postos ORDER BY nome_posto').fetchall()
    conn.close()
    return render_template('postos.html', postos=lista, user=current_user())


# ------------------------
# Usuários (Gerencial) - owner only
# ------------------------

@app.route('/gerencial/usuarios', methods=['GET', 'POST'])
@login_required
@owner_required
def usuarios():
    conn = get_db_connection()

    if request.method == 'POST':
        action = (request.form.get('action') or 'create').strip().lower()

        if action == 'create':
            username = (request.form.get('username') or '').strip().lower()
            senha = request.form.get('senha') or ''
            role = (request.form.get('role') or 'manager').strip()
            posto_id = request.form.get('posto_id')
            posto_id = int(posto_id) if posto_id else None

            if username and senha and role in ('owner', 'manager'):
                conn.execute(
                    'INSERT INTO users (username, password_hash, role, posto_id, created_at) VALUES (?,?,?,?,?)',
                    (username, generate_password_hash(senha), role, posto_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S')),
                )
                conn.commit()

        elif action == 'update':
            user_id = request.form.get('user_id')
            user_id = int(user_id) if user_id else None
            username = (request.form.get('username') or '').strip().lower()
            role = (request.form.get('role') or 'manager').strip()
            posto_id = request.form.get('posto_id')
            posto_id = int(posto_id) if posto_id else None
            active = 1 if (request.form.get('active') == '1') else 0
            senha = request.form.get('senha') or ''

            if user_id and username and role in ('owner', 'manager'):
                # Impede "manager" sem posto
                if role == 'manager' and not posto_id:
                    posto_id = None

                conn.execute(
                    'UPDATE users SET username = ?, role = ?, posto_id = ?, active = ? WHERE id = ?',
                    (username, role, posto_id, active, user_id),
                )
                if senha.strip():
                    conn.execute(
                        'UPDATE users SET password_hash = ? WHERE id = ?',
                        (generate_password_hash(senha), user_id),
                    )
                conn.commit()

        elif action == 'toggle':
            user_id = request.form.get('user_id')
            user_id = int(user_id) if user_id else None
            if user_id:
                row = conn.execute('SELECT active FROM users WHERE id = ?', (user_id,)).fetchone()
                if row:
                    new_active = 0 if int(row['active'] or 0) == 1 else 1
                    conn.execute('UPDATE users SET active = ? WHERE id = ?', (new_active, user_id))
                    conn.commit()

        return redirect(url_for('usuarios'))

    users = conn.execute(
        'SELECT u.*, p.nome_posto FROM users u LEFT JOIN postos p ON p.id = u.posto_id ORDER BY u.id DESC'
    ).fetchall()
    postos = conn.execute('SELECT * FROM postos ORDER BY nome_posto').fetchall()
    conn.close()
    return render_template(
        'usuarios.html',
        users=users,
        postos=postos,
        user=current_user(),
    )



@app.route('/logout')
def logout():
    """Encerra sessão (admin/gerente) e volta para seleção de posto."""
    session.clear()
    return redirect(url_for('selecionar_posto'))

if __name__ == '__main__':
    init_db()
    app.run(host='0.0.0.0', port=int(os.getenv('PORT', '5000')))

@app.errorhandler(Exception)
def handle_exception(e):
    import traceback
    return f'<pre>{traceback.format_exc()}</pre>', 500
